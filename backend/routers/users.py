from __future__ import annotations

import re
import secrets
from io import BytesIO
from html import escape

import pandas as pd
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func, or_
from sqlalchemy.orm import Session, selectinload

from ..dependencies import get_current_user, get_db, require_roles
from ..models import ActionItem, Department, EmailTemplate, Meeting, Notice, Opportunity, Result, ScheduledEmail, User, UserRole, meeting_attendees, parent_student_links, teacher_student_links, user_departments
from ..schemas import ParentLinkCreate, StudentRegistration, TeacherLinkCreate, UserAdminUpdate, UserCreate, UserUpdateSettings
from ..services.auth_service import get_password_hash
from ..services.email_service import send_plain_email
from ..services.serialization import serialize_user


router = APIRouter(prefix='/api/users', tags=['users'])


def _valid_password(password: str) -> bool:
    return len(password) >= 12 and any(char.islower() for char in password) and any(char.isupper() for char in password) and any(char.isdigit() for char in password) and any(not char.isalnum() for char in password)


def _valid_email(email: str) -> bool:
    return bool(re.fullmatch(r'[^\s@]+@[^\s@]+\.[^\s@]+', email))


def _generated_password() -> str:
    # token_urlsafe supplies high entropy; the suffix guarantees every password rule.
    return f'{secrets.token_urlsafe(12)}Aa1!'


def _queue_welcome_email(tasks: BackgroundTasks, user: User, password: str):
    subject = 'Welcome to The Bridge School Portal'
    plain_body = f'''Dear {user.name},

Your account has been created on The Bridge School Portal.

Email: {user.email}
Password: {password}
Role: {user.role.value.title()}

Login at: http://localhost:5173

Please change your password after logging in for the first time.

Regards,
The Bridge School'''
    html_body = f'''<div style="font-family:Arial,sans-serif;color:#1a1a1a;max-width:600px;margin:auto;padding:28px">
      <h2 style="color:#1B2B6B;margin:0 0 20px">The Bridge School</h2>
      <p>Dear <strong>{escape(user.name)}</strong>,</p>
      <p>Your account has been created on The Bridge School Portal. Here are your login credentials:</p>
      <table style="border-collapse:collapse;width:100%;background:#f7f8fc;border:1px solid #e2e5f0"><tr><td style="padding:12px;color:#6b7280">Email</td><td style="padding:12px;font-weight:600">{escape(user.email)}</td></tr><tr><td style="padding:12px;color:#6b7280">Password</td><td style="padding:12px;font-weight:600">{escape(password)}</td></tr><tr><td style="padding:12px;color:#6b7280">Role</td><td style="padding:12px;font-weight:600">{escape(user.role.value.title())}</td></tr></table>
      <p style="margin-top:20px">For your security, please change your password after you sign in.</p></div>'''
    tasks.add_task(send_plain_email, user.email, subject, plain_body, html_body=html_body)


def _delete_user_record(db: Session, user: User) -> None:
    """Delete a user and owned/dependent records explicitly so removal works on all supported databases."""
    user_id = user.id
    guardian_ids = [guardian.id for guardian in user.guardians] if user.role == UserRole.student else []
    db.execute(delete(parent_student_links).where(or_(parent_student_links.c.parent_id == user_id, parent_student_links.c.student_id == user_id)))
    db.execute(delete(teacher_student_links).where(or_(teacher_student_links.c.teacher_id == user_id, teacher_student_links.c.student_id == user_id)))
    db.execute(delete(meeting_attendees).where(meeting_attendees.c.user_id == user_id))
    db.execute(delete(user_departments).where(user_departments.c.user_id == user_id))
    db.query(ActionItem).filter(ActionItem.assigned_to == user_id).delete(synchronize_session=False)
    db.query(Result).filter(or_(Result.student_id == user_id, Result.uploaded_by == user_id)).delete(synchronize_session=False)
    db.query(ScheduledEmail).filter(ScheduledEmail.created_by == user_id).delete(synchronize_session=False)
    db.query(EmailTemplate).filter(EmailTemplate.created_by == user_id).delete(synchronize_session=False)
    db.query(Notice).filter(Notice.created_by == user_id).delete(synchronize_session=False)
    db.query(Opportunity).filter(Opportunity.created_by == user_id).delete(synchronize_session=False)
    db.query(Meeting).filter(Meeting.created_by == user_id).delete(synchronize_session=False)
    db.query(User).filter(User.id == user_id).delete(synchronize_session=False)
    db.flush()
    # Removing a student also removes guardians who no longer belong to any student.
    for guardian_id in guardian_ids:
        has_children = db.query(parent_student_links).filter(parent_student_links.c.parent_id == guardian_id).first()
        if not has_children:
            guardian = db.get(User, guardian_id)
            if guardian:
                _delete_user_record(db, guardian)


def _invalid_family_users(db: Session) -> list[User]:
    users = db.query(User).options(selectinload(User.children), selectinload(User.guardians)).filter(User.role.in_([UserRole.student, UserRole.parent])).all()
    return [user for user in users if (user.role == UserRole.student and not user.guardians) or (user.role == UserRole.parent and not user.children)]


@router.post('/cleanup-invalid-family-records')
def cleanup_invalid_family_records(db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    invalid_users = _invalid_family_users(db)
    removed = [{'id': user.id, 'name': user.name, 'role': user.role.value} for user in invalid_users]
    for user in invalid_users:
        _delete_user_record(db, user)
    db.commit()
    return {'removed': removed, 'count': len(removed)}


IMPORT_TEMPLATES = {
    'students': {
        'filename': 'students-and-guardians-import-template.xlsx',
        'sheet_title': 'Students & Guardians',
        'columns': ['student_name', 'student_email', 'student_password', 'guardian_1_name', 'guardian_1_email', 'guardian_2_name', 'guardian_2_email'],
        'required': {'student_name', 'student_email', 'guardian_1_name', 'guardian_1_email'},
        'instructions': [
            'One student per row. At least one parent or guardian is required.',
            'Guardian 2 is optional. The same guardian email can be used for siblings.',
            'Leave password columns blank to generate a secure temporary password automatically.',
            'Any supplied password must have 12+ characters with upper/lowercase, a number, and a symbol.',
        ],
    },
    'teachers': {
        'filename': 'teachers-import-template.xlsx',
        'sheet_title': 'Teachers',
        'columns': ['teacher_name', 'teacher_email', 'teacher_password', 'head_teacher'],
        'required': {'teacher_name', 'teacher_email'},
        'instructions': [
            'One teacher per row.',
            'Use Yes or No in head_teacher. Only use Yes for a head teacher.',
            'Leave teacher_password blank to generate a secure temporary password automatically.',
        ],
    },
    'staff': {
        'filename': 'staff-import-template.xlsx',
        'sheet_title': 'Staff',
        'columns': ['staff_name', 'staff_email', 'staff_password', 'department'],
        'required': {'staff_name', 'staff_email', 'department'},
        'instructions': [
            'One staff member per row. Department is required.',
            'A new department/domain is created automatically if it does not yet exist.',
            'Leave staff_password blank to generate a secure temporary password automatically.',
        ],
    },
}

IMPORT_PREFIXES = {'students': 'student', 'teachers': 'teacher', 'staff': 'staff'}


def _excel_template(template: dict) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = template['sheet_title']
    sheet.append(template['columns'])
    header_fill = PatternFill('solid', fgColor='1B2B6B')
    for column_index, column in enumerate(template['columns'], start=1):
        cell = sheet.cell(1, column_index)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        sheet.column_dimensions[get_column_letter(column_index)].width = max(16, len(column) + 3)
    sheet.freeze_panes = 'A2'
    instructions = workbook.create_sheet('Instructions')
    instructions.column_dimensions['A'].width = 110
    instructions['A1'] = f"{template['sheet_title']} Import Template"
    instructions['A1'].font = Font(bold=True, size=14, color='1B2B6B')
    for row_number, instruction in enumerate(template['instructions'], start=3):
        instructions.cell(row_number, 1, instruction).alignment = Alignment(wrap_text=True, vertical='top')
    instructions['A8'] = 'Required columns'
    instructions['A8'].font = Font(bold=True, color='1B2B6B')
    instructions['A9'] = ', '.join(sorted(template['required']))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{template["filename"]}"'})


@router.get('/templates/{import_type}.xlsx')
def users_template(import_type: str):
    template = IMPORT_TEMPLATES.get(import_type)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='Unknown import template')
    return _excel_template(template)


@router.get('')
def list_users(db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    users = db.query(User).options(selectinload(User.departments), selectinload(User.children), selectinload(User.guardians), selectinload(User.teachers), selectinload(User.students_taught)).order_by(User.created_at.desc()).all()
    return [serialize_user(user) for user in users]


@router.post('')
def create_user(payload: UserCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    email = payload.email.strip().lower()
    if not payload.name.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Name is required')
    if not _valid_email(email):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Enter a valid email address')
    if not _valid_password(payload.password):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Temporary password must be 12+ characters with uppercase, lowercase, a number, and a symbol')
    if db.query(User).filter(func.lower(User.email) == email).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Email already exists')
    try:
        role = UserRole(payload.role)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Invalid role') from exc
    if role == UserRole.student:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Register students through the Student & guardian form so every student has a linked parent or guardian')
    department_name = payload.department.strip() if payload.department else None
    if role == UserRole.staff and not department_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Department/domain is required for staff')
    user = User(
        name=payload.name.strip(),
        email=email,
        hashed_password=get_password_hash(payload.password),
        role=role,
        head_teacher=payload.head_teacher,
        department=department_name if role == UserRole.staff else None,
        is_active=payload.is_active,
    )
    db.add(user)
    if department_name and role == UserRole.staff:
        department = db.query(Department).filter(Department.name == department_name).first()
        if not department:
            department = Department(name=department_name)
            db.add(department)
        user.departments.append(department)
    db.commit()
    db.refresh(user)

    _queue_welcome_email(background_tasks, user, payload.password)
    response = serialize_user(user)
    response['invitation_queued'] = True
    return response


@router.post('/students')
def create_student(payload: StudentRegistration, background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    student_email = payload.email.strip().lower()
    if not _valid_email(student_email) or not payload.name.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Enter a student name and valid email address')
    if not _valid_password(payload.password):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Student password must be 12+ characters with uppercase, lowercase, a number, and a symbol')
    if db.query(User).filter(func.lower(User.email) == student_email).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Email already exists')

    guardian_emails = set()
    resolved_guardians: list[tuple[User, str | None]] = []
    for guardian_data in payload.guardians:
        email = guardian_data.email.strip().lower()
        if not guardian_data.name.strip() or not _valid_email(email):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Each guardian needs a name and valid email address')
        if email in guardian_emails:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Each guardian email can only be used once')
        guardian_emails.add(email)
        existing = db.query(User).filter(func.lower(User.email) == email).first()
        if existing and existing.role != UserRole.parent:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f'{email} belongs to an existing {existing.role.value} account and cannot be used as a guardian')
        if existing:
            resolved_guardians.append((existing, None))
        else:
            password = _generated_password()
            guardian = User(name=guardian_data.name.strip(), email=email, hashed_password=get_password_hash(password), role=UserRole.parent, is_active=True)
            db.add(guardian)
            resolved_guardians.append((guardian, password))

    student = User(name=payload.name.strip(), email=student_email, hashed_password=get_password_hash(payload.password), role=UserRole.student, is_active=True)
    db.add(student)
    db.flush()
    for guardian, _ in resolved_guardians:
        student.guardians.append(guardian)
    db.commit()
    db.refresh(student)
    _queue_welcome_email(background_tasks, student, payload.password)
    for guardian, password in resolved_guardians:
        if password:
            _queue_welcome_email(background_tasks, guardian, password)
    response = serialize_user(student)
    response['invitation_queued'] = True
    response['created_guardian_accounts'] = sum(1 for _, password in resolved_guardians if password)
    return response


def _normalise_import_rows(frame: pd.DataFrame) -> list[dict[str, str]]:
    frame.columns = [str(column).strip().lower().replace(' ', '_') for column in frame.columns]
    return [{key: str(value).strip() for key, value in row.items()} for row in frame.to_dict(orient='records')]


def _validate_import_password(value: str, label: str) -> str | None:
    if not value:
        return None
    if not _valid_password(value):
        return f'{label} must be 12+ characters and include uppercase, lowercase, a number, and a symbol'
    return None


@router.post('/import')
async def import_users(
    background_tasks: BackgroundTasks,
    import_type: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.admin)),
):
    """Import a role-specific Excel workbook only after the complete batch has passed validation."""
    template = IMPORT_TEMPLATES.get(import_type)
    if not template:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Choose Students & Guardians, Teachers, or Staff before importing')
    filename = (file.filename or '').lower()
    if not filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Upload an Excel .xlsx or .xls file')
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='The import file must be 10 MB or smaller')
    try:
        frame = pd.read_excel(BytesIO(contents), dtype=str).fillna('')
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='The Excel file could not be read') from exc
    rows = _normalise_import_rows(frame)
    missing = sorted(template['required'] - set(frame.columns))
    if missing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"This is not the {template['sheet_title']} template. Missing columns: {', '.join(missing)}")
    if frame.empty:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='The workbook has no user rows')

    errors: list[dict] = []
    primary_emails: set[str] = set()
    guardian_details: dict[str, str] = {}
    for index, row in enumerate(rows, start=2):
        prefix = IMPORT_PREFIXES[import_type]
        name = row.get(f'{prefix}_name', '')
        email = row.get(f'{prefix}_email', '').lower()
        password = row.get(f'{prefix}_password', '')
        if not name or not _valid_email(email):
            errors.append({'row': index, 'message': 'A name and valid email address are required'})
        password_error = _validate_import_password(password, f'{prefix.title()} password')
        if password_error:
            errors.append({'row': index, 'message': password_error})
        if email in primary_emails:
            errors.append({'row': index, 'message': 'This email is repeated in the workbook'})
        primary_emails.add(email)
        if email and db.query(User.id).filter(func.lower(User.email) == email).first():
            errors.append({'row': index, 'message': 'An account with this email already exists'})
        if import_type == 'staff' and not row.get('department', ''):
            errors.append({'row': index, 'message': 'Department is required for every staff member'})
        if import_type == 'teachers' and row.get('head_teacher', '').lower() not in {'', 'yes', 'no', 'true', 'false'}:
            errors.append({'row': index, 'message': 'head_teacher must be Yes or No'})
        if import_type == 'students':
            for number in (1, 2):
                guardian_name = row.get(f'guardian_{number}_name', '')
                guardian_email = row.get(f'guardian_{number}_email', '').lower()
                if number == 1 and (not guardian_name or not _valid_email(guardian_email)):
                    errors.append({'row': index, 'message': 'Guardian 1 needs a name and valid email address'})
                    continue
                if guardian_email and (not guardian_name or not _valid_email(guardian_email)):
                    errors.append({'row': index, 'message': f'Guardian {number} needs both a name and valid email address'})
                    continue
                if guardian_email:
                    if guardian_email == email:
                        errors.append({'row': index, 'message': 'A student and guardian cannot use the same email address'})
                    previous_name = guardian_details.get(guardian_email)
                    if previous_name and previous_name.lower() != guardian_name.lower():
                        errors.append({'row': index, 'message': 'The same guardian email has different names in the workbook'})
                    guardian_details[guardian_email] = guardian_name
                    existing = db.query(User).filter(func.lower(User.email) == guardian_email).first()
                    if existing and existing.role != UserRole.parent:
                        errors.append({'row': index, 'message': f'{guardian_email} belongs to a non-parent account'})
    for guardian_email in guardian_details:
        if guardian_email in primary_emails:
            errors.append({'row': 'Workbook', 'message': f'{guardian_email} is used as both a primary account and guardian'})
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={'message': 'No users were imported. Correct the listed rows and try again.', 'errors': errors})

    created: list[tuple[User, str]] = []
    try:
        for row in rows:
            prefix = IMPORT_PREFIXES[import_type]
            email = row[f'{prefix}_email'].lower()
            password = row.get(f'{prefix}_password') or _generated_password()
            if import_type != 'students':
                role = UserRole.teacher if import_type == 'teachers' else UserRole.staff
                user = User(
                    name=row[f'{prefix}_name'], email=email, hashed_password=get_password_hash(password), role=role,
                    head_teacher=import_type == 'teachers' and row.get('head_teacher', '').lower() in {'yes', 'true'},
                    department=row.get('department') or None, is_active=True,
                )
                db.add(user)
                if role == UserRole.staff:
                    department = db.query(Department).filter(Department.name == row['department']).first() or Department(name=row['department'])
                    db.add(department)
                    user.departments.append(department)
                created.append((user, password))
                continue
            student = User(name=row['student_name'], email=email, hashed_password=get_password_hash(password), role=UserRole.student, is_active=True)
            db.add(student)
            db.flush()
            created.append((student, password))
            for number in (1, 2):
                guardian_email = row.get(f'guardian_{number}_email', '').lower()
                guardian_name = row.get(f'guardian_{number}_name', '')
                if not guardian_email:
                    continue
                guardian = db.query(User).filter(func.lower(User.email) == guardian_email).first()
                if guardian and guardian.role != UserRole.parent:
                    raise ValueError(f'{guardian_email} belongs to a non-parent account')
                if not guardian:
                    guardian_password = _generated_password()
                    guardian = User(name=guardian_name, email=guardian_email, hashed_password=get_password_hash(guardian_password), role=UserRole.parent, is_active=True)
                    db.add(guardian)
                    created.append((guardian, guardian_password))
                if guardian not in student.guardians:
                    student.guardians.append(guardian)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f'No users were imported: {str(exc)}') from exc
    for user, password in created:
        _queue_welcome_email(background_tasks, user, password)
    return {'rows_imported': len(rows), 'accounts_created': len(created), 'message': f'{len(rows)} {template["sheet_title"].lower()} row(s) imported. Credential emails have been queued.'}


@router.patch('/{user_id}')
def update_user(user_id: int, payload: UserAdminUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='User not found')
    if 'name' in payload.model_fields_set:
        user.name = payload.name.strip() if payload.name else user.name
    if 'email' in payload.model_fields_set:
        email = payload.email.strip().lower() if payload.email else ''
        if not _valid_email(email):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Enter a valid email address')
        conflict = db.query(User).filter(func.lower(User.email) == email, User.id != user.id).first()
        if conflict:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Email already exists')
        user.email = email
    if 'role' in payload.model_fields_set and payload.role:
        try:
            new_role = UserRole(payload.role)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Invalid role') from exc
        if user.role == UserRole.parent and new_role != UserRole.parent and user.children:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='This parent is linked to students. Remove those links before changing their role.')
        if user.role == UserRole.student and new_role != UserRole.student and user.guardians:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='This student is linked to guardians. Remove those links before changing their role.')
        if new_role == UserRole.student and user.role != UserRole.student:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Students must be registered through the Student & guardian form so a parent or guardian is created and linked at the same time.')
        user.role = new_role
    if 'head_teacher' in payload.model_fields_set:
        user.head_teacher = bool(payload.head_teacher)
    if 'is_active' in payload.model_fields_set:
        user.is_active = bool(payload.is_active)
    if 'department' in payload.model_fields_set:
        user.department = payload.department.strip() if payload.department else None
    db.commit()
    db.refresh(user)
    return serialize_user(user)


@router.post('/{student_id}/guardians')
def link_guardian(student_id: int, payload: ParentLinkCreate, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    student = db.get(User, student_id)
    parent = db.get(User, payload.parent_id)
    if not student or student.role != UserRole.student:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Student not found')
    if not parent or parent.role != UserRole.parent:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Parent or guardian account not found')
    if parent not in student.guardians:
        if len(student.guardians) >= 2:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='A student can have a maximum of two parents or guardians')
        student.guardians.append(parent)
        db.commit()
    db.refresh(student)
    return serialize_user(student)


@router.delete('/{student_id}/guardians/{parent_id}')
def unlink_guardian(student_id: int, parent_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    student = db.get(User, student_id)
    parent = db.get(User, parent_id)
    if not student or student.role != UserRole.student or not parent or parent.role != UserRole.parent:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='Student or guardian not found')
    if len(student.guardians) <= 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='A student must always have at least one linked parent or guardian')
    if parent in student.guardians:
        student.guardians.remove(parent)
        db.commit()
    db.refresh(student)
    return serialize_user(student)


@router.post('/{student_id}/teachers')
def link_teacher(student_id: int, payload: TeacherLinkCreate, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    student = db.get(User, student_id)
    teacher = db.get(User, payload.teacher_id)
    if not student or student.role != UserRole.student:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Student not found')
    if not teacher or teacher.role != UserRole.teacher:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Teacher account not found')
    if teacher not in student.teachers:
        student.teachers.append(teacher)
        db.commit()
    db.refresh(student)
    return serialize_user(student)


@router.delete('/{student_id}/teachers/{teacher_id}')
def unlink_teacher(student_id: int, teacher_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    student = db.get(User, student_id)
    teacher = db.get(User, teacher_id)
    if not student or student.role != UserRole.student or not teacher or teacher.role != UserRole.teacher:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='Student or teacher not found')
    if teacher in student.teachers:
        student.teachers.remove(teacher)
        db.commit()
    db.refresh(student)
    return serialize_user(student)


@router.delete('/{user_id}')
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_roles(UserRole.admin))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='User not found')
    if user.role == UserRole.parent:
        for student in user.children:
            if len(student.guardians) <= 1:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f'Cannot remove {user.name}: {student.name} would be left without a parent or guardian')
    _delete_user_record(db, user)
    db.commit()
    return {'detail': 'User deleted'}


@router.patch('/me/settings')
def update_my_settings(payload: UserUpdateSettings, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Name cannot be empty')
        current_user.name = name
    if payload.profile_picture_url is not None:
        picture_url = payload.profile_picture_url or None
        # The browser sends an image data URL.  Keep a practical upper bound so
        # a mistaken upload cannot fill the database, while allowing normal
        # profile photos (and clients that do not pre-compress them).
        if picture_url is not None and len(picture_url) > 5_000_000:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Profile picture is too large to save')
        current_user.profile_picture_url = picture_url
    if payload.email_notifications_enabled is not None:
        current_user.email_notifications_enabled = payload.email_notifications_enabled
    db.commit()
    db.refresh(current_user)
    return serialize_user(current_user)
